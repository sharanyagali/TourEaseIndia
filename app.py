from flask import Flask, render_template, request, redirect, url_for, session, g, jsonify, send_file
import sqlite3
import os
from functools import wraps
from datetime import datetime
import json
import uuid
import csv
import io
import threading
import hashlib
import urllib.request
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, PageBreak, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False
try:
    from PIL import Image as PILImage
    PIL_AVAILABLE = True
except Exception:
    PIL_AVAILABLE = False
try:
    import firebase_admin
    from firebase_admin import credentials, auth as fb_auth, firestore
    FIREBASE_AVAILABLE = True
except Exception:
    FIREBASE_AVAILABLE = False

app = Flask(__name__)
app.secret_key = 'your_secret_key'
DATABASE = os.path.join(os.path.dirname(__file__), 'database.db')
FIREBASE_ENABLED = False
FS = None
ACTIVE_USERS = 0
TOTAL_USERS = 0
PDF_DIR = os.path.join(os.path.dirname(__file__), 'static', 'pdfs')

def init_firebase():
    global FIREBASE_ENABLED, FS
    if not FIREBASE_AVAILABLE:
        FIREBASE_ENABLED = False
        return
    if not firebase_admin._apps:
        cred_path = os.environ.get('FIREBASE_CREDENTIALS')
        if cred_path and os.path.exists(cred_path):
            cred = credentials.Certificate(cred_path)
            firebase_admin.initialize_app(cred)
            FS = firestore.client()
            FIREBASE_ENABLED = True
        else:
            FIREBASE_ENABLED = False

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
        db.row_factory = sqlite3.Row
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db:
        db.close()

def init_db():
    db = get_db()
    with open('schema.sql', 'r', encoding='utf-8') as f:
        db.executescript(f.read())
    db.commit()

def ensure_columns():
    db = get_db()
    cols = db.execute("PRAGMA table_info(destinations)").fetchall()
    names = {c['name'] for c in cols}
    if 'lat' not in names:
        db.execute("ALTER TABLE destinations ADD COLUMN lat REAL")
    if 'lng' not in names:
        db.execute("ALTER TABLE destinations ADD COLUMN lng REAL")
    db.execute("UPDATE destinations SET lat=?, lng=? WHERE name LIKE '%Goa%'", (15.3, 73.8))
    db.execute("UPDATE destinations SET lat=?, lng=? WHERE name LIKE '%Shimla%'", (31.10, 77.17))
    db.execute("UPDATE destinations SET lat=?, lng=? WHERE name LIKE '%Munnar%'", (10.09, 77.06))
    db.execute("UPDATE destinations SET lat=?, lng=? WHERE name LIKE '%Hyderabad%'", (17.3850, 78.4746))
    db.execute("UPDATE destinations SET lat=?, lng=? WHERE name LIKE '%Delhi%'", (28.5275, 77.0689))
    db.commit()
def login_required(f=None):
    # Support both @login_required and @login_required()
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if 'uid' not in session:
                return redirect(url_for('login'))
            return fn(*args, **kwargs)
        return wrapper
    if f is None:
        return decorator
    return decorator(f)

@app.route('/')
def index():
    if 'uid' not in session:
        return redirect(url_for('login'))
    db = get_db()
    destinations = db.execute("SELECT * FROM destinations WHERE name != 'Shimla' ORDER BY name").fetchall()
    all_places = _read_places_csv()
    states = sorted({p['state'] for p in all_places if p['state']})
    return render_template('index.html', destinations=destinations, states=states)

@app.route('/login', methods=['GET'])
def login():
    if 'uid' in session:
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'GET':
        if 'uid' in session:
            return redirect(url_for('index'))
        return render_template('register.html')
    name = (request.form.get('name') or '').strip()
    email = (request.form.get('email') or '').strip()
    phone = (request.form.get('phone') or '').strip()
    password = (request.form.get('password') or '').strip()
    
    # Validation
    if not email:
        return render_template('register.html', error='Email is required', name=name, phone=phone)
    if not password or len(password) < 6:
        return render_template('register.html', error='Password must be at least 6 characters long', name=name, phone=phone)
    
    db = get_db()
    row = db.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
    if row:
        return render_template('register.html', error='Email already registered. Please login.', name=name, phone=phone)
    
    db.execute("INSERT INTO users (name, email, phone, password) VALUES (?, ?, ?, ?)", (name or email, email, phone, password))
    db.commit()
    uid = db.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()['id']
    session['uid'] = uid
    session['user_email'] = email
    session['user_name'] = name or email
    global ACTIVE_USERS, TOTAL_USERS
    ACTIVE_USERS += 1
    TOTAL_USERS += 1
    init_firebase()
    if FIREBASE_ENABLED:
        try:
            FS.collection('users').document(str(uid)).set({
                'uid': str(uid),
                'name': name or email,
                'email': email,
                'phone': phone,
                'isOnline': True,
                'registered_at': firestore.SERVER_TIMESTAMP
            }, merge=True)
            act_ref = FS.collection('user_activity').document()
            act_ref.set({
                'uid': str(uid),
                'login_ts': datetime.utcnow(),
                'logout_ts': None
            })
            session['activity_id'] = act_ref.id
        except Exception:
            pass
    return redirect(url_for('index'))
@app.route('/logout')
def logout():
    uid = session.get('uid')
    activity_id = session.get('activity_id')
    global ACTIVE_USERS
    if FIREBASE_ENABLED and uid:
        try:
            if activity_id:
                FS.collection('user_activity').document(activity_id).update({
                    'logout_ts': datetime.utcnow()
                })
            FS.collection('users').document(uid).set({'isOnline': False}, merge=True)
        except Exception:
            pass
    if ACTIVE_USERS > 0:
        ACTIVE_USERS -= 1
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html', user_name=session.get('user_name', session.get('user_email')))

@app.route('/season/<season>')
@login_required
def season(season):
    db = get_db()
    packages = db.execute("SELECT * FROM packages WHERE season=?", (season,)).fetchall()
    return render_template('season.html', packages=packages, season=season)

def _infer_best(text):
    t = (text or '').lower()
    if 'winter' in t: return 'Winter'
    if 'monsoon' in t or 'rain' in t: return 'Monsoon'
    if 'summer' in t or 'hot' in t: return 'Summer'
    return 'All'

def _normalize_state(state):
    s = (state or '').strip()
    sl = s.lower()
    if not s:
        return s
    an_variants = {
        'anamanda nicobar islands',
        'andaman nicobar islands',
        'andaman & nicobar islands',
        'andamans and nicobar',
        'andaman and nicobar',
        'andaman and nicobar island',
        'andaman nicobar',
    }
    if sl in an_variants:
        return 'Andaman and Nicobar Islands'
    return s

def _force_telangana_if_hyderabad(state, city):
    c = (city or '').strip()
    if 'hyderabad' in c.lower():
        return 'Telangana'
    return state
def _transport_for_state(state):
    s = (state or '').strip()
    metros = {
        'Telangana': [{'label':'Hyderabad Metro','url':'https://www.ltmetro.com/'}],
        'Delhi': [{'label':'Delhi Metro','url':'https://www.delhimetrorail.com/'}],
        'Maharashtra': [{'label':'Mumbai Metro','url':'https://www.mmrda.maharashtra.gov.in/metro'} , {'label':'Pune Metro','url':'https://www.punemetrorail.org/'}],
        'Karnataka': [{'label':'Bengaluru Metro','url':'https://english.bmrc.co.in/'}],
        'Tamil Nadu': [{'label':'Chennai Metro','url':'https://chennaimetrorail.org/'}],
        'Gujarat': [{'label':'Ahmedabad Metro','url':'https://www.gujaratmetrorail.com/'}],
        'West Bengal': [{'label':'Kolkata Metro','url':'https://mtp.indianrailways.gov.in/'}],
        'Uttar Pradesh': [{'label':'Lucknow Metro','url':'https://www.lmrcl.com/'} , {'label':'Noida Metro','url':'https://www.nmrcnoida.com/'}],
        'Rajasthan': [{'label':'Jaipur Metro','url':'https://www.jaipurmetrorail.in/'}],
        'Kerala': [{'label':'Kochi Metro','url':'https://kochimetro.org/'}],
        'Haryana': [],
    }
    return {
        'rapido': {'label':'Rapido','url':'https://www.rapido.bike/'},
        'bus': {'label':'Bus Tickets','url':'https://www.redbus.in/'},
        'metros': metros.get(s, []),
        'has_metro': bool(metros.get(s))
    }
def _read_places_csv():

    csv_path = os.path.join(os.path.dirname(__file__), 'Top Indian Places to Visit.csv')
    if not os.path.exists(csv_path):
        return []
    places = []
    image_map = _read_places_excel()
    dedup = {}
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for r in reader:
            name = (r.get('Name') or '').strip()
            city = (r.get('City') or '').strip()
            if 'qutub minar' in name.lower():
                continue
            state_raw = (r.get('State') or '').strip()
            if state_raw.isdigit():
                continue
            state = _normalize_state(state_raw)
            state = _force_telangana_if_hyderabad(state, city)
            fee = (r.get('Entrance Fee in INR') or '').strip()
            if 'taj mahal' in name.lower():
                fee = 'Indians ₹50 · Foreigners ₹1100'
            rating = (r.get('Google review rating') or '').strip()
            time_hrs = (r.get('time needed to visit in hrs') or '').strip()
            weekly_off = (r.get('Weekly Off') or '').strip()
            dslr = (r.get('DSLR Allowed') or '').strip()
            best_text = (r.get('Best Time to visit') or '').strip()
            best = _infer_best(best_text)
            key = f"{city}|{name}".lower()
            rec = {
                'name': name,
                'city': city,
                'state': state,
                'fee': fee,
                'rating': rating,
                'time_hrs': time_hrs,
                'weekly_off': weekly_off,
                'dslr': dslr,
                'best_text': best_text,
                'best': best,
                'map_q': (name + ', ' + city + ', ' + state).strip(', '),
                'image': image_map.get(f"{state}|{city}|{name}".lower()) or "https://upload.wikimedia.org/wikipedia/commons/6/6e/India_collage.jpg"
            }
            prev = dedup.get(key)
            if not prev:
                dedup[key] = rec
            else:
                def rv(x):
                    try:
                        return float(x['rating']) if x['rating'] else 0.0
                    except Exception:
                        return 0.0
                # Prefer Telangana when city Hyderabad, else higher rating
                if 'hyderabad' in city.lower():
                    dedup[key] = rec
                elif rv(rec) > rv(prev):
                    dedup[key] = rec
    places = list(dedup.values())
    def fallback_image_for(name):
        n = (name or '').lower()
        if 'taj mahal' in n:
            return 'https://cdn.britannica.com/86/170586-050-AB7FEFAE/Taj-Mahal-Agra-India.jpg'
        if 'ram' in n and 'mandir' in n:
            return 'https://upload.wikimedia.org/wikipedia/commons/thumb/7/72/Ayodhya_Ram_Mandir.jpg/640px-Ayodhya_Ram_Mandir.jpg'
        if 'qutub minar' in n:
            return 'https://images.mapsofworld.com/allwonders/qutub-minar.jpg'
        return None
    def _enrich_place_info(r):
        n = (r.get('name') or '').lower()
        c = (r.get('city') or '').lower()
        s = (r.get('state') or '').lower()
        if 'qutub minar' in n:
            r['image'] = 'https://images.mapsofworld.com/allwonders/qutub-minar.jpg'
            r['best'] = r.get('best') or 'Winter'
            r['best_text'] = r.get('best_text') or 'Outdoor site; mild weather Nov–Feb; evening light views'
            r['desc'] = 'UNESCO tower showcasing Indo-Islamic architecture in Delhi.'
        elif 'taj mahal' in n:
            r['image'] = 'https://cdn.britannica.com/86/170586-050-AB7FEFAE/Taj-Mahal-Agra-India.jpg'
            r['desc'] = 'Marble mausoleum on the Yamuna; sunrise reflections in Agra.'
            r['fee'] = 'Indians ₹50 · Foreigners ₹1100'
        elif 'ram' in n and 'mandir' in n:
            r['desc'] = 'Temple complex in Ayodhya; festival ambience and devotional music.'
        elif 'munnar' in n:
            r['best'] = r.get('best') or 'Winter'
            r['best_text'] = r.get('best_text') or 'Tea estates and viewpoints; best Sep–May'
            r['desc'] = 'Tea gardens, mist, winding roads, and high-altitude vistas.'
        elif 'gangtok' in n:
            r['best'] = r.get('best') or 'All'
            r['best_text'] = r.get('best_text') or 'All-year; clear spring/autumn views of Kanchenjunga'
            r['desc'] = 'Cloud-wrapped capital of Sikkim with mountain panoramas.'
        elif 'kashmir' in n or 'srinagar' in c:
            r['best'] = r.get('best') or 'Summer'
            r['best_text'] = r.get('best_text') or 'Pleasant spring/summer; autumn colours Sep–Nov'
            r['desc'] = 'Valley known as Paradise on Earth with lakes and gardens.'
        elif 'manali' in n:
            r['best'] = r.get('best') or 'Summer'
            r['best_text'] = r.get('best_text') or 'Best Oct–Jun; snow in winter, cool summers'
            r['desc'] = 'Himalayan hill station with valleys, deodar forests, and snow peaks.'
        elif 'jaipur' in n:
            r['best'] = r.get('best') or 'Winter'
            r['best_text'] = r.get('best_text') or 'Pleasant Oct–Mar; forts, palaces, bazaars'
            r['desc'] = 'Pink City famed for heritage architecture and vibrant markets.'
        return r
    def add_if_missing(city, name, state, best, best_text, rating, time_hrs, weekly_off):
        key = f"{city}|{name}".lower()
        if key not in dedup:
            img = image_map.get(f"{state}|{city}|{name}".lower()) or fallback_image_for(name)
            rec = {
                'name': name,
                'city': city,
                'state': state,
                'fee': '—',
                'rating': rating,
                'time_hrs': time_hrs,
                'weekly_off': weekly_off,
                'dslr': '—',
                'best_text': best_text,
                'best': best,
                'map_q': (name + ', ' + city + ', ' + state).strip(', '),
                'image': img
            }
            rec = _enrich_place_info(rec)
            dedup[key] = rec
    add_if_missing('Agra', 'Taj Mahal', 'Uttar Pradesh', 'Winter', 'Pleasant weather; sunrise views; avoid summer heat', '4.8', '2.0', 'Friday')
    add_if_missing('Ayodhya', 'Shri Ram Janmabhoomi Mandir', 'Uttar Pradesh', 'Winter', 'Festivals; mild weather; book stay in advance', '4.7', '2.0', '—')
    add_if_missing('New Delhi', 'Qutub Minar', 'Delhi', 'Winter', 'Outdoor site; mild weather Nov–Feb; evening light views', '4.7', '1.5', '—')
    add_if_missing('New Delhi', 'Taj Mahal', 'Delhi', 'Winter', 'Pleasant weather; sunrise views; avoid summer heat', '4.8', '2.0', 'Friday')
    places = [ _enrich_place_info(p) for p in dedup.values() ]
    return places

def _read_places_excel():
    xlsx_path = os.path.join(os.path.dirname(__file__), 'Data set of famous India tourist places along with there images.xlsx')
    if not os.path.exists(xlsx_path) or not OPENPYXL_AVAILABLE:
        return {}
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        headers = {}
        image_col = None
        state_col = None
        city_col = None
        name_col = None
        for j, cell in enumerate(ws[1], start=1):
            h = (str(cell.value) or '').strip()
            headers[h] = j
        # Heuristic header mapping
        def find(*candidates):
            for c in candidates:
                if c in headers:
                    return headers[c]
            return None
        image_col = find('Image', 'Image URL', 'Image Link', 'image', 'image_url', 'image link')
        state_col = find('State', 'state')
        city_col = find('City', 'city')
        name_col = find('Name', 'Place', 'name')
        if not (state_col and city_col and name_col and image_col):
            return {}
        out = {}
        for i in range(2, ws.max_row + 1):
            s = str(ws.cell(row=i, column=state_col).value or '').strip()
            c = str(ws.cell(row=i, column=city_col).value or '').strip()
            n = str(ws.cell(row=i, column=name_col).value or '').strip()
            img = str(ws.cell(row=i, column=image_col).value or '').strip()
            if not (s and n):
                continue
            key = f"{s}|{c}|{n}".lower()
            if img:
                out[key] = img
        return out
    except Exception:
        return {}

def _ensure_min_places_for_all_states(all_places, season):
    states = sorted({p['state'] for p in all_places if p['state']})
    def match(p):
        t = (p.get('best_text') or '')
        return (p.get('best') == season) or (season.lower() in t.lower())
    by_state = {}
    for s in states:
        by_state[s] = [p for p in all_places if p['state'] == s and match(p)]
    def rv(x):
        try:
            return float(x['rating']) if x['rating'] else 0.0
        except Exception:
            return 0.0
    out = []
    for s in states:
        selected = by_state.get(s, [])
        if len(selected) < 4:
            pool = [p for p in all_places if p['state'] == s]
            pool_sorted = sorted(pool, key=lambda x: (rv(x), x['name']), reverse=True)
            used = {(p['city'], p['name']) for p in selected}
            for p in pool_sorted:
                key = (p['city'], p['name'])
                if key not in used:
                    q = dict(p)
                    q['best'] = season
                    q['best_text'] = f'Best in {season}'
                    q['_fallback'] = True
                    selected.append(q)
                    used.add(key)
                if len(selected) >= 4:
                    break
        out.extend(selected)
    return out

@app.route('/season_places')
@login_required
def season_places():
    season = (request.args.get('season') or 'All').title()
    all_places = _read_places_csv()
    if season == 'All':
        filtered = all_places
    else:
        filtered = _ensure_min_places_for_all_states(all_places, season)
    states = sorted({p['state'] for p in all_places if p['state']})
    sel_state = request.args.get('state')
    if sel_state:
        filtered = [p for p in filtered if p['state'] == sel_state]
    transport = _transport_for_state(sel_state) if sel_state else None
    return render_template('season.html', places=filtered, season=season, states=states, state=sel_state or '', transport=transport)

@app.route('/places_by_state')
@login_required
def places_by_state():
    all_places = _read_places_csv()
    states = sorted({p['state'] for p in all_places if p['state']})
    state = (request.args.get('state') or (states[0] if states else '')).strip()
    season = (request.args.get('season') or 'All').title()
    filtered = [p for p in all_places if p['state'] == state]
    if season != 'All':
        filtered = [p for p in filtered if (p['best']==season or season.lower() in (p['best_text'] or '').lower())]
    if len(filtered) < 4:
        pool = [p for p in all_places if p['state'] == state]
        def rv(x):
            try:
                return float(x['rating']) if x['rating'] else 0.0
            except Exception:
                return 0.0
        pool_sorted = sorted(pool, key=lambda x: (rv(x), x['name']), reverse=True)
        existing = { (p['city'], p['name']) for p in filtered }
        for p in pool_sorted:
            key = (p['city'], p['name'])
            if key not in existing:
                filtered.append(p)
                existing.add(key)
            if len(filtered) >= 4:
                break
    transport = _transport_for_state(state)
    return render_template('season.html', places=filtered, season=season, states=states, state=state, transport=transport)

@app.route('/states')
@login_required
def states():
    all_places = _read_places_csv()
    by_state = {}
    for p in all_places:
        s = p['state'] or 'Unknown'
        if s not in by_state:
            by_state[s] = {'state': s, 'count': 0, 'ratings': [], 'season_counts': {'Winter':0,'Monsoon':0,'Summer':0,'All':0}, 'image': None, 'places': []}
        by_state[s]['count'] += 1
        try:
            r = float(p['rating']) if p['rating'] else None
            if r is not None:
                by_state[s]['ratings'].append(r)
        except Exception:
            pass
        by_state[s]['season_counts'][p['best']] = by_state[s]['season_counts'].get(p['best'],0) + 1
        if not by_state[s]['image'] and p.get('image'):
            by_state[s]['image'] = p['image']
        by_state[s]['places'].append({'name': p['name'], 'city': p['city'], 'image': p.get('image'), 'rating': p.get('rating'), 'best': p.get('best')})
    states_data = []
    for s, v in by_state.items():
        avg = round(sum(v['ratings'])/len(v['ratings']), 2) if v['ratings'] else None
        dom = max(v['season_counts'], key=lambda k: v['season_counts'][k]) if v['season_counts'] else 'All'
        # sort places by rating desc (fallback to name)
        def rating_val(x):
            try:
                return float(x['rating']) if x['rating'] else 0.0
            except Exception:
                return 0.0
        places_sorted = sorted(v['places'], key=lambda x: (rating_val(x), x['name']), reverse=True)
        top_places = places_sorted[:6]
        rep_image = v['image'] or (top_places[0]['image'] if (top_places and top_places[0]['image']) else None)
        states_data.append({'state': s, 'count': v['count'], 'avg_rating': avg, 'dominant_season': dom, 'image': rep_image, 'top_places': top_places})
    states_data.sort(key=lambda x: x['state'])
    return render_template('states.html', states=states_data)

@app.route('/place/<int:dest_id>')
@login_required
def place(dest_id):
    db = get_db()
    destination = db.execute("SELECT * FROM destinations WHERE id=?", (dest_id,)).fetchone()
    hotels = db.execute("SELECT * FROM hotels WHERE destination_id=?", (dest_id,)).fetchall()
    foods = db.execute("SELECT * FROM food WHERE destination_id=?", (dest_id,)).fetchall()
    transport = _transport_for_state(destination['state']) if destination else None
    return render_template('place.html', destination=destination, hotels=hotels, foods=foods, transport=transport)

@app.route('/offline_guide', methods=['GET', 'POST'])
@login_required
def offline_guide():
    db = get_db()
    destinations = db.execute("SELECT * FROM destinations WHERE name != 'Shimla' ORDER BY name").fetchall()
    if request.method == 'POST':
        dest_id = int(request.form.get('destination_id'))
        return redirect(url_for('place', dest_id=dest_id))
    all_places = _read_places_csv()
    states = sorted({p['state'] for p in all_places if p['state']})
    return render_template('offline_guide.html', destinations=destinations, states=states, places=all_places)

@app.route('/maps', methods=['GET'])
@login_required
def maps_routes():
    db = get_db()
    destinations = db.execute("SELECT * FROM destinations WHERE name != 'Shimla' ORDER BY name").fetchall()
    dest_id = request.args.get('destination_id', type=int)
    selected = None
    if dest_id:
        selected = db.execute("SELECT * FROM destinations WHERE id=?", (dest_id,)).fetchone()
    all_places = _read_places_csv()
    states = sorted({p['state'] for p in all_places if p['state']})
    sel_state = (request.args.get('state') or '').strip()
    return render_template('maps_routes.html', destinations=destinations, selected=selected, states=states, sel_state=sel_state)

@app.route('/tickets', methods=['GET'])
@login_required
def tickets():
    from_state = (request.args.get('from') or '').strip()
    to_state = (request.args.get('to') or '').strip()
    date = (request.args.get('date') or '').strip()
    def qurl(base, q):
        import urllib.parse
        return base + urllib.parse.quote(q)
    flights_url = qurl("https://www.google.com/search?q=", f"flights from {from_state} to {to_state} {date}") if from_state and to_state else "https://www.google.com/travel/flights"
    trains_url = qurl("https://www.google.com/search?q=", f"IRCTC trains from {from_state} to {to_state} {date}") if from_state and to_state else "https://www.irctc.co.in/nget/train-search"
    buses_url = qurl("https://www.google.com/search?q=", f"bus tickets from {from_state} to {to_state} {date}") if from_state and to_state else "https://www.redbus.in/"
    return render_template('tickets.html',
                           from_state=from_state,
                           to_state=to_state,
                           date=date,
                           flights_url=flights_url,
                           trains_url=trains_url,
                           buses_url=buses_url)

@app.route('/itinerary/hyderabad')
@login_required
def itinerary_hyderabad():
    days = request.args.get('days', type=int)
    if not days or days < 3 or days > 7:
        days = 7
    def image_for(name):
        if 'Golconda' in name: return "https://thumbs.dreamstime.com/b/golconda-fort-3720711.jpg"
        if 'Hussain Sagar' in name: return "https://upload.wikimedia.org/wikipedia/commons/thumb/f/f2/Hussain_Sagar.jpg/640px-Hussain_Sagar.jpg"
        if 'Charminar' in name: return "https://upload.wikimedia.org/wikipedia/commons/thumb/d/d9/Charminar_%28Night_View%29.jpg/640px-Charminar_%28Night_View%29.jpg"
        if 'Birla Mandir' in name: return "https://hyderabadtourism.co.in/wp-content/uploads/2024/05/Birla_Mandir_Hyderabad_2-1024x649.jpg"
        if 'Jagannath' in name: return "https://upload.wikimedia.org/wikipedia/commons/thumb/5/58/Hyd_Jagannath_Temple.jpg/640px-Hyd_Jagannath_Temple.jpg"
        if 'Qutb Shahi' in name: return "https://upload.wikimedia.org/wikipedia/commons/thumb/1/1a/Qutb_Shahi_Tombs_Complex.jpg/640px-Qutb_Shahi_Tombs_Complex.jpg"
        if 'Shilparamam' in name: return "https://upload.wikimedia.org/wikipedia/commons/thumb/f/f1/Shilparamam_Village.JPG/640px-Shilparamam_Village.JPG"
        if 'Salar Jung' in name: return "https://upload.wikimedia.org/wikipedia/commons/thumb/6/6d/Salar_Jung_Museum.jpg/640px-Salar_Jung_Museum.jpg"
        if 'Nizam Museum' in name: return "https://upload.wikimedia.org/wikipedia/commons/thumb/7/76/Nizam_Museum.jpg/640px-Nizam_Museum.jpg"
        if 'Necklace Road' in name: return "https://upload.wikimedia.org/wikipedia/commons/thumb/5/55/Necklace_Road.jpg/640px-Necklace_Road.jpg"
        if 'Ramoji' in name: return "https://letsgohyderabad.com/wp-content/uploads/2024/12/Ramoji_Film_City-1.jpg"
        if 'Laad Bazaar' in name: return "https://upload.wikimedia.org/wikipedia/commons/thumb/f/f1/Laad_Bazaar%2C_Hyderabad.jpg/640px-Laad_Bazaar%2C_Hyderabad.jpg"
        if 'Abids' in name: return "https://upload.wikimedia.org/wikipedia/commons/thumb/d/db/Abids%2C_Hyderabad_3.jpg/640px-Abids%2C_Hyderabad_3.jpg"
        return "https://upload.wikimedia.org/wikipedia/commons/6/6e/India_collage.jpg"
    places = [
        {'name':'Golconda Fort','when':'Morning','schedule':'9:00 AM – 12:30 PM','transport':'Hotel vehicle','guide':'Available','food':'Water, snacks nearby','shopping':'','highlight':'Citadel, acoustics','lat':17.3833,'lng':78.4011},
        {'name':'Hussain Sagar Lake','when':'Night','schedule':'7:00 PM – 9:00 PM','transport':'Hotel vehicle or RTC bus','guide':'','food':'Street food at Necklace Road','shopping':'','highlight':'Fountain show, lighting','lat':17.4239,'lng':78.4738},
        {'name':'Charminar','when':'Evening','schedule':'5:30 PM – 8:30 PM','transport':'Metro/auto or hotel','guide':'','food':'Irani chai, Osmania biscuits, biryani','shopping':'Laad Bazaar','highlight':'Monument lights','lat':17.3616,'lng':78.4747},
        {'name':'Birla Mandir','when':'Morning','schedule':'7:00 AM – 9:00 AM','transport':'RTC bus','guide':'','food':'Nearby cafes','shopping':'','highlight':'Marble temple views','lat':17.4062,'lng':78.4691},
        {'name':'Jagannath Temple','when':'Morning','schedule':'9:00 AM – 10:30 AM','transport':'RTC bus','guide':'','food':'','shopping':'','highlight':'Architecture','lat':17.4392,'lng':78.3980},
        {'name':'Qutb Shahi Tombs','when':'Flexible','schedule':'','transport':'Auto/hotel','guide':'','food':'','shopping':'','highlight':'Architecture & gardens','lat':17.3947,'lng':78.3966},
        {'name':'Shilparamam','when':'Flexible','schedule':'','transport':'Auto/hotel','guide':'','food':'Local snacks','shopping':'Crafts','highlight':'Performances','lat':17.4698,'lng':78.3860},
        {'name':'Salar Jung Museum','when':'Midday','schedule':'11:00 AM – 2:00 PM','transport':'Auto/hotel','guide':'','food':'','shopping':'','highlight':'Collections','lat':17.3711,'lng':78.4809},
        {'name':'Nizam Museum','when':'Afternoon','schedule':'2:30 PM – 4:00 PM','transport':'Auto/hotel','guide':'','food':'','shopping':'','highlight':'Royal artifacts','lat':17.3698,'lng':78.4723},
        {'name':'Necklace Road','when':'Evening','schedule':'6:00 PM – 8:00 PM','transport':'Auto/hotel','guide':'','food':'Street food','shopping':'','highlight':'Sunset stroll','lat':17.4231,'lng':78.4656},
        {'name':'Ramoji Film City','when':'Full Day','schedule':'8:30 AM – 5:30 PM','transport':'Hotel vehicle','guide':'Available','food':'Park canteens','shopping':'','highlight':'Shows and sets','lat':17.2567,'lng':78.6800},
        {'name':'Laad Bazaar','when':'Afternoon','schedule':'12:30 PM – 2:30 PM','transport':'Auto/hotel','guide':'','food':'Snacks nearby','shopping':'Bangles & crafts','highlight':'Markets','lat':17.3612,'lng':78.4740},
        {'name':'Abids','when':'Afternoon','schedule':'3:00 PM – 5:00 PM','transport':'Auto/hotel','guide':'','food':'City cafes','shopping':'Clothing','highlight':'Shopping avenues','lat':17.3950,'lng':78.4740},
    ]
    for p in places:
        p['image'] = image_for(p['name'])
    order = [
        ['Golconda Fort','Hussain Sagar Lake','Charminar'],
        ['Birla Mandir','Jagannath Temple'],
        ['Qutb Shahi Tombs','Shilparamam'],
        ['Salar Jung Museum','Nizam Museum','Necklace Road'],
        ['Ramoji Film City'],
        ['Laad Bazaar','Abids'],
        ['Relax & Departure']
    ]
    def by_name(n):
        for p in places:
            if p['name'] == n:
                return p
        if n == 'Relax & Departure':
            return {'name':n,'when':'','schedule':'','transport':'Hotel or metro to connect','guide':'','food':'Breakfast','shopping':'','highlight':'','image':image_for(n),'lat':None,'lng':None}
        return {'name':n}
    base_days = [ [by_name(n) for n in grp] for grp in order ]
    if days == 7:
        program = base_days
    elif days == 6:
        program = base_days[:5] + [base_days[5] + base_days[6]]
    elif days == 5:
        program = [base_days[0], base_days[1], base_days[2] + base_days[3], base_days[4], base_days[5] + base_days[6]]
    elif days == 4:
        program = [base_days[0], base_days[1] + base_days[2], base_days[3] + base_days[4], base_days[5] + base_days[6]]
    else:
        program = [base_days[0], base_days[1] + base_days[3], base_days[4] + base_days[5] + base_days[6]]
    return render_template('itinerary_hyderabad.html', days=days, program=program, state='Telangana', cover_image="https://upload.wikimedia.org/wikipedia/commons/thumb/d/d9/Charminar_%28Night_View%29.jpg/640px-Charminar_%28Night_View%29.jpg")

@app.route('/itinerary')
@login_required
def itinerary_state():
    state = (request.args.get('state') or '').strip()
    days = request.args.get('days', type=int)
    if not days or days < 3 or days > 7:
        days = 5
    all_places = _read_places_csv()
    places = [p for p in all_places if p['state'] == state] if state else []
    def rv(x):
        try:
            return float(x['rating']) if x['rating'] else 0.0
        except Exception:
            return 0.0
    places_sorted = sorted(places, key=lambda x: (rv(x), x['name']), reverse=True)
    count = min(len(places_sorted), days * 3) if places_sorted else 0
    selected = places_sorted[:count]
    program = []
    when_labels = ['Morning', 'Afternoon', 'Evening']
    schedules = ['9:00 AM – 12:00 PM', '1:30 PM – 4:30 PM', '6:00 PM – 8:30 PM']
    has_metro = bool(_transport_for_state(state).get('metros'))
    transport_text = ('Metro/auto or bus' if has_metro else 'Auto or RTC bus')
    day = []
    for idx, p in enumerate(selected):
        slot = idx % 3
        day.append({
            'name': p['name'],
            'city': p['city'],
            'when': when_labels[slot],
            'schedule': schedules[slot],
            'transport': transport_text,
            'guide': '',
            'food': 'Local snacks & cafés',
            'shopping': '',
            'highlight': p.get('best_text') or '',
            'lat': None,
            'lng': None,
            'image': p.get('image'),
        })
        if len(day) == 3:
            program.append(day)
            day = []
    if day:
        program.append(day)
    cover_image = selected[0].get('image') if selected else None
    return render_template('itinerary_hyderabad.html', days=days, program=program, state=state or 'India', cover_image=cover_image)

def _pdf_filename(state, season):
    s = (state or 'All').split('?')[0].split('&')[0].strip().replace(' ', '-')
    se = (_sanitize_season(season) or 'All').replace(' ', '-')
    return f"{s}_{se}.pdf"

def _pdf_path(state, season):
    if not os.path.exists(PDF_DIR):
        os.makedirs(PDF_DIR, exist_ok=True)
    return os.path.join(PDF_DIR, _pdf_filename(state, season))

def _sanitize_season(season_raw):
    t = (season_raw or 'All')
    t = t.split('?')[0].split('&')[0].strip().lower()
    if 'winter' in t: return 'Winter'
    if 'monsoon' in t or 'rain' in t: return 'Monsoon'
    if 'summer' in t or 'hot' in t: return 'Summer'
    return 'All'

def _clean_param(val):
    return (val or '').split('?')[0].split('&')[0].strip()

def _fetch_and_compress(url, max_w=640, quality=60):
    if not url:
        return None
    try:
        data = urllib.request.urlopen(url, timeout=4).read()
        bio = io.BytesIO(data)
        if PIL_AVAILABLE:
            img = PILImage.open(bio)
            if img.mode in ('RGBA','P'):
                img = img.convert('RGB')
            w, h = img.size
            if w > max_w:
                ratio = max_w / float(w)
                img = img.resize((int(w*ratio), int(h*ratio)))
            out = io.BytesIO()
            img.save(out, format='JPEG', quality=quality, optimize=True)
            out.seek(0)
            return out
        bio.seek(0)
        return bio
    except Exception:
        return None

def _find_nearby_attractions(state, city, exclude_name='', limit=4):
    """Find nearby attractions in the same state or city"""
    all_places = _read_places_csv()
    def rv(x):
        try:
            return float(x['rating']) if x['rating'] else 0.0
        except Exception:
            return 0.0
    
    # First try to find attractions in the same city
    same_city = [p for p in all_places if p['state']==state and p['city']==city and p['name'] != exclude_name]
    if same_city:
        same_city = sorted(same_city, key=lambda x: (rv(x), x['name']), reverse=True)
        return same_city[:limit]
    
    # If not enough in same city, get from same state
    same_state = [p for p in all_places if p['state']==state and p['name'] != exclude_name]
    if same_state:
        same_state = sorted(same_state, key=lambda x: (rv(x), x['name']), reverse=True)
        return same_state[:limit]
    
    return []

def _build_pdf(state, season, path):
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError('reportlab not available')
    all_places = _read_places_csv()
    if season and season != 'All':
        def match(p):
            t = (p.get('best_text') or '')
            return (p.get('best') == season) or (season.lower() in t.lower())
        all_places = [p for p in all_places if match(p)]
    if state:
        all_places = [p for p in all_places if p['state'] == state]
    def rv(x):
        try:
            return float(x['rating']) if x['rating'] else 0.0
        except Exception:
            return 0.0
    places = sorted(all_places, key=lambda x: (rv(x), x['name']), reverse=True)[:20]
    if not places and state:
        pool = _read_places_csv()
        pool = [p for p in pool if p['state'] == state]
        places = sorted(pool, key=lambda x: (rv(x), x['name']), reverse=True)[:12]
    styles = getSampleStyleSheet()
    story = []
    title = Paragraph(f"🇮🇳 TourEase India Travel Guide", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 8))
    subtitle = Paragraph(f"<b>{state or 'All States'}</b> · {season or 'All Seasons'}", styles['Heading3'])
    story.append(subtitle)
    story.append(Spacer(1, 12))
    meta = Table([[f"📍 Top Places: {len(places)}", f"📅 Generated: {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')}"]], colWidths=[3*inch, 3*inch])
    meta.setStyle(TableStyle([('GRID',(0,0),(-1,-1),1,colors.HexColor('#e2e8f0')),('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#f0f7ff'))]))
    story.append(meta)
    story.append(Spacer(1, 16))
    if not places:
        story.append(Paragraph("No data available for this selection.", styles['Normal']))
        doc = SimpleDocTemplate(path, pagesize=A4)
        doc.build(story)
        return
    for idx, p in enumerate(places):
        # Main place heading
        heading = Paragraph(f"<b>{idx+1}. {p['name']}</b>", styles['Heading2'])
        story.append(heading)
        location = Paragraph(f"<i>{p['city']} · {p['state']}</i>", styles['Normal'])
        story.append(location)
        story.append(Spacer(1, 6))
        
        # Details table
        grid = [
            [Paragraph("<b>Best Season</b>", styles['Normal']), Paragraph(p.get('best') or 'All', styles['Normal'])],
            [Paragraph("<b>Entrance Fee</b>", styles['Normal']), Paragraph(p.get('fee') or '—', styles['Normal'])],
            [Paragraph("<b>Google Rating</b>", styles['Normal']), Paragraph(p.get('rating') or '—', styles['Normal'])],
            [Paragraph("<b>Time Needed</b>", styles['Normal']), Paragraph((p.get('time_hrs') and (str(p.get('time_hrs'))+' hours')) or '—', styles['Normal'])],
            [Paragraph("<b>Weekly Off</b>", styles['Normal']), Paragraph(p.get('weekly_off') or 'Check', styles['Normal'])],
            [Paragraph("<b>DSLR Allowed</b>", styles['Normal']), Paragraph(p.get('dslr') or 'Check', styles['Normal'])],
        ]
        tbl = Table(grid, colWidths=[1.8*inch, 4.2*inch])
        tbl.setStyle(TableStyle([
            ('BOX',(0,0),(-1,-1),1,colors.HexColor('#cbd5e1')),
            ('INNERGRID',(0,0),(-1,-1),0.5,colors.HexColor('#e2e8f0')),
            ('BACKGROUND',(0,0),(0,-1),colors.HexColor('#f0f7ff')),
            ('VALIGN',(0,0),(-1,-1),'TOP'),
            ('LEFTPADDING',(0,0),(-1,-1),8),
            ('RIGHTPADDING',(0,0),(-1,-1),8),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 8))
        
        # Main image
        img_stream = _fetch_and_compress(p.get('image'), max_w=500, quality=70)
        if img_stream:
            story.append(RLImage(img_stream, width=5.0*inch, height=3.5*inch))
            story.append(Spacer(1, 8))
        
        # Nearby attractions section
        nearby = _find_nearby_attractions(state or p['state'], p.get('city'), p['name'], limit=3)
        if nearby:
            story.append(Paragraph("<b>🎯 Nearby Attractions (Worth Visiting):</b>", styles['Heading3']))
            story.append(Spacer(1, 6))
            nearby_table_data = []
            for n in nearby:
                try:
                    rating_str = f"★ {n.get('rating', 'N/A')}" if n.get('rating') else "★ Rated"
                    nearby_table_data.append([
                        Paragraph(f"<b>{n['name']}</b><br/>{n.get('city', '')}", styles['Normal']),
                        Paragraph(f"Fee: {n.get('fee', '—')}<br/>{rating_str}", styles['Normal']),
                    ])
                except Exception:
                    pass
            
            if nearby_table_data:
                nearby_tbl = Table(nearby_table_data, colWidths=[3.5*inch, 2.5*inch])
                nearby_tbl.setStyle(TableStyle([
                    ('BOX',(0,0),(-1,-1),1,colors.HexColor('#dcfce7')),
                    ('INNERGRID',(0,0),(-1,-1),0.5,colors.HexColor('#bbf7d0')),
                    ('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#f0fdf4')),
                    ('LEFTPADDING',(0,0),(-1,-1),8),
                    ('RIGHTPADDING',(0,0),(-1,-1),8),
                ]))
                story.append(nearby_tbl)
                story.append(Spacer(1, 12))
        
        # Add page break between places for better readability (every 2 places)
        if (idx + 1) % 2 == 0 and idx < len(places) - 1:
            story.append(PageBreak())
        else:
            story.append(Spacer(1, 20))
    
    # Footer
    story.append(Spacer(1, 20))
    story.append(Paragraph("<hr/>", styles['Normal']))
    story.append(Paragraph("✈️ Happy Traveling with TourEase India! 🇮🇳", styles['Normal']))
    story.append(Paragraph(f"Downloaded on {datetime.utcnow().strftime('%d %B %Y at %H:%M UTC')}", styles['Normal']))
    
    doc = SimpleDocTemplate(path, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    doc.build(story)

def _build_itinerary_pdf(state, days):
    prog = []
    all_places = _read_places_csv()
    sel = [p for p in all_places if p['state'] == state] if state else []
    def rv(x):
        try:
            return float(x['rating']) if x['rating'] else 0.0
        except Exception:
            return 0.0
    places_sorted = sorted(sel, key=lambda x: (rv(x), x['name']), reverse=True)
    count = min(len(places_sorted), (days or 5) * 3) if places_sorted else 0
    selected = places_sorted[:count]
    when_labels = ['Morning', 'Afternoon', 'Evening']
    schedules = ['9:00 AM – 12:00 PM', '1:30 PM – 4:30 PM', '6:00 PM – 8:30 PM']
    has_metro = bool(_transport_for_state(state).get('metros'))
    transport_text = ('Metro/auto or bus' if has_metro else 'Auto or RTC bus')
    day = []
    for idx, p in enumerate(selected):
        slot = idx % 3
        day.append({
            'name': p['name'],
            'city': p['city'],
            'when': when_labels[slot],
            'schedule': schedules[slot],
            'transport': transport_text,
            'highlight': p.get('best_text') or '',
            'image': p.get('image'),
        })
        if len(day) == 3:
            prog.append(day); day = []
    if day:
        prog.append(day)
    bio = io.BytesIO()
    if REPORTLAB_AVAILABLE:
        styles = getSampleStyleSheet()
        story = [Paragraph(f"{state or 'India'} · {days or 5}-Day Itinerary", styles['Title']), Spacer(1,12)]
        for d_idx, day_items in enumerate(prog, start=1):
            story.append(Paragraph(f"Day {d_idx}", styles['Heading2']))
            for item in day_items:
                story.append(Paragraph(f"{item['name']} · {item['schedule']} · {item['when']}", styles['Normal']))
                story.append(Paragraph(f"Transport: {item['transport']}", styles['Normal']))
                if item['highlight']:
                    story.append(Paragraph(f"Highlights: {item['highlight']}", styles['Normal']))
                img_stream = _fetch_and_compress(item.get('image'))
                if img_stream:
                    story.append(Spacer(1,6))
                    story.append(RLImage(img_stream, width=4.5*inch, height=3.0*inch))
                story.append(Spacer(1,10))
            story.append(Spacer(1,12))
        doc = SimpleDocTemplate(bio, pagesize=A4)
        doc.build(story)
        bio.seek(0)
        return bio
    else:
        lines = [f"{state or 'India'} · {days or 5}-Day Itinerary", ""]
        for d_idx, day_items in enumerate(prog, start=1):
            lines.append(f"Day {d_idx}")
            for item in day_items:
                lines.append(f"{item['name']} · {item['schedule']} · {item['when']} · Transport: {item['transport']}")
            lines.append("")
        y_start = 800; step = 16
        stream = ["BT /F1 12 Tf 50 800 Td"]
        def _sanitize_text(s):
            t = (s or '')
            t = t.replace('–','-').replace('—','-').replace('’',"'").replace('‘',"'").replace('“','"').replace('”','"').replace('…','...').replace('•','-')
            return t
        def esc(s):
            t = _sanitize_text(s)
            return (t or '').replace('\\','\\\\').replace('(','\\(').replace(')','\\)')
        for ln in lines:
            stream.append(f"({esc(ln)}) Tj 0 -{step} Td")
        stream.append("ET")
        sd = ("\n".join(stream)).encode('latin-1', errors='replace')
        objs = [
            b"<< /Type /Catalog /Pages 2 0 R >>",
            b"<< /Type /Pages /Count 1 /Kids [3 0 R] >>",
            b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 5 0 R >> >> /Contents 6 0 R >>",
            b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
            b"<< /Length " + str(len(sd)).encode('latin-1') + b" >>\nstream\n" + sd + b"\nendstream",
        ]
        out = io.BytesIO(); out.write(b"%PDF-1.4\n")
        offsets = [0]
        for i,o in enumerate(objs, start=1):
            offsets.append(out.tell()); out.write(f"{i} 0 obj\n".encode('latin-1', errors='replace')); out.write(o); out.write(b"\nendobj\n")
        xref = out.tell(); out.write(f"xref\n0 {len(objs)+1}\n".encode('latin-1', errors='replace')); out.write(b"0000000000 65535 f \n")
        for off in offsets[1:]:
            out.write(f"{off:010d} 00000 n \n".encode('latin-1', errors='replace'))
        out.write(b"trailer\n"); out.write(f"<< /Size {len(objs)+1} /Root 1 0 R >>\n".encode('latin-1', errors='replace'))
        out.write(f"startxref\n{xref}\n%%EOF".encode('latin-1', errors='replace'))
        out.seek(0)
        return out

@app.route('/download_itinerary', methods=['GET'])
@login_required
def download_itinerary():
    state = _clean_param(request.args.get('state') or '')
    days = request.args.get('days', type=int) or 5
    bio = _build_itinerary_pdf(state, days)
    return send_file(bio, as_attachment=True, download_name=f"TourEase_Itinerary_{state or 'India'}_{days}days.pdf", mimetype='application/pdf')

def _itinerary_pdf_path(state, days):
    if not os.path.exists(PDF_DIR):
        os.makedirs(PDF_DIR, exist_ok=True)
    s = (state or 'India').strip().replace(' ', '-')
    return os.path.join(PDF_DIR, f"TourEase_Itinerary_{s}_{days}days.pdf")

def _build_itinerary_pdf_file(state, days, path):
    if REPORTLAB_AVAILABLE:
        styles = getSampleStyleSheet()
        story = [Paragraph(f"{state or 'India'} · {days or 5}-Day Travel Itinerary", styles['Title']), Spacer(1,16)]
        prog_bio = _build_itinerary_pdf(state, days)
        # If _build_itinerary_pdf produced a reportlab doc in memory, rebuild content simply:
        all_places = _read_places_csv()
        sel = [p for p in all_places if p['state'] == state] if state else []
        if not sel:
            story.append(Paragraph("No data available for this selection.", styles['Normal']))
        else:
            # Simple per-day render
            def rv(x):
                try: return float(x['rating']) if x['rating'] else 0.0
                except: return 0.0
            places_sorted = sorted(sel, key=lambda x: (rv(x), x['name']), reverse=True)
            count = min(len(places_sorted), (days or 5) * 3)
            selected = places_sorted[:count]
            when_labels = ['Morning','Afternoon','Evening']
            schedules = ['9:00 AM – 12:00 PM','1:30 PM – 4:30 PM','6:00 PM – 8:30 PM']
            has_metro = bool(_transport_for_state(state).get('metros'))
            transport_text = ('Metro/auto or bus' if has_metro else 'Auto or RTC bus')
            # Render
            for d_idx in range(1, (days or 5)+1):
                story.append(Paragraph(f"Day {d_idx}", styles['Heading2']))
                day_items = selected[(d_idx-1)*3 : d_idx*3]
                for i, p in enumerate(day_items):
                    slot = i
                    info_rows = [
                        [Paragraph("Place", styles['Normal']), Paragraph(f"{p['name']} · {p['city']}, {p['state']}", styles['Normal'])],
                        [Paragraph("Schedule", styles['Normal']), Paragraph(f"{schedules[slot]} · {when_labels[slot]}", styles['Normal'])],
                        [Paragraph("Transport", styles['Normal']), Paragraph(transport_text, styles['Normal'])],
                        [Paragraph("Best Season", styles['Normal']), Paragraph(p.get('best') or 'All', styles['Normal'])],
                        [Paragraph("Rating", styles['Normal']), Paragraph(p.get('rating') or '—', styles['Normal'])],
                        [Paragraph("Entrance Fee", styles['Normal']), Paragraph(p.get('fee') or '—', styles['Normal'])],
                        [Paragraph("Season Note", styles['Normal']), Paragraph(p.get('best_text') or '—', styles['Normal'])],
                    ]
                    tbl = Table(info_rows, colWidths=[1.6*inch, 4.4*inch])
                    tbl.setStyle(TableStyle([('GRID',(0,0),(-1,-1),0.5,colors.lightgrey),('BACKGROUND',(0,0),(-1,-1),colors.whitesmoke)]))
                    story.append(tbl)
                    img_stream = _fetch_and_compress(p.get('image'))
                    if img_stream:
                        story.append(Spacer(1,6))
                        story.append(RLImage(img_stream, width=4.5*inch, height=3.0*inch))
                    story.append(Spacer(1,10))
        doc = SimpleDocTemplate(path, pagesize=A4)
        doc.build(story)
        return
    # Fallback: use in-memory builder and persist to file
    bio = _build_itinerary_pdf(state, days)
    try:
        # Write to file for caching
        with open(path, 'wb') as f:
            f.write(bio.getvalue())
    except Exception:
        pass

@app.route('/download_itinerary_pdf', methods=['GET'])
@login_required
def download_itinerary_pdf():
    state = _clean_param(request.args.get('state') or '')
    days = request.args.get('days', type=int) or 5
    path = _itinerary_pdf_path(state, days)
    if os.path.exists(path):
        return send_file(path, as_attachment=True, download_name=os.path.basename(path), mimetype='application/pdf', conditional=True)
    _build_itinerary_pdf_file(state, days, path)
    if os.path.exists(path):
        return send_file(path, as_attachment=True, download_name=os.path.basename(path), mimetype='application/pdf', conditional=True)
    # Last resort: in-memory
    bio = _build_itinerary_pdf(state, days)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f"TourEase_Itinerary_{state or 'India'}_{days}days.pdf", mimetype='application/pdf')
def _ensure_pdf_async(state, season, path):
    def run():
        try:
            _build_pdf(state, season, path)
        except Exception:
            pass
    t = threading.Thread(target=run, daemon=True)
    t.start()

def _build_basic_pdf(state, season, path):
    all_places = _read_places_csv()
    if season and season != 'All':
        def match(p):
            t = (p.get('best_text') or '')
            return (p.get('best') == season) or (season.lower() in t.lower())
        all_places = [p for p in all_places if match(p)]
    if state:
        all_places = [p for p in all_places if p['state'] == state]
    def rv(x):
        try:
            return float(x['rating']) if x['rating'] else 0.0
        except Exception:
            return 0.0
    places = sorted(all_places, key=lambda x: (rv(x), x['name']), reverse=True)[:30]
    def _sanitize_text(s):
        t = (s or '')
        t = t.replace('–','-').replace('—','-').replace('’',"'").replace('‘',"'").replace('“','"').replace('”','"').replace('…','...').replace('•','-')
        return t
    def esc(s):
        t = _sanitize_text(s)
        return (t or '').replace('\\','\\\\').replace('(','\\(').replace(')','\\)')
    lines = []
    lines.append(f"TourEase India · {state or 'All States'} · {season or 'All'}")
    lines.append(f"Generated: {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')}")
    for p in places:
        lines.append(f"{p['name']} — {p['city']}, {p['state']}")
        lines.append(f"Best: {p.get('best') or 'All'} · Rating: {p.get('rating') or '—'} · Fee: {p.get('fee') or '—'}")
        lines.append(f"Time: {(p.get('time_hrs') and (str(p.get('time_hrs'))+' hrs')) or '—'} · Weekly Off: {p.get('weekly_off') or '—'}")
        lines.append(f"Note: {p.get('best_text') or '—'}")
        lines.append("")
    pages = []
    y_start = 800
    step = 16
    max_lines = int(y_start/step) - 10
    chunk = []
    for i, ln in enumerate(lines):
        chunk.append(ln)
        if len(chunk) >= max_lines:
            pages.append(chunk)
            chunk = []
    if chunk:
        pages.append(chunk)
    objs = []
    def add_obj(s):
        if isinstance(s, bytes):
            objs.append(s)
        else:
            objs.append(str(s).encode('latin-1', errors='replace'))
    add_obj(b"<< /Type /Catalog /Pages 2 0 R >>")
    kids_refs = []
    font_obj_num = 5
    content_start_num = 6
    page_objs = []
    contents_objs = []
    for idx, chunk in enumerate(pages):
        y = y_start
        stream_lines = []
        stream_lines.append("BT /F1 12 Tf 50 800 Td")
        for ln in chunk:
            stream_lines.append(f"({esc(ln)}) Tj 0 -{step} Td")
        stream_lines.append("ET")
        stream_data = ("\n".join(stream_lines)).encode('latin-1', errors='replace')
        contents_objs.append(stream_data)
        pg_num = 3 + idx
        cnt_num = content_start_num + idx
        kids_refs.append(f"{pg_num} 0 R")
        page_objs.append(f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 {font_obj_num} 0 R >> >> /Contents {cnt_num} 0 R >>")
    add_obj(f"<< /Type /Pages /Count {len(pages)} /Kids [{' '.join(kids_refs)}] >>")
    for pdef in page_objs:
        add_obj(pdef)
    add_obj(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    for sd in contents_objs:
        add_obj(b"<< /Length " + str(len(sd)).encode('latin-1') + b" >>\nstream\n" + sd + b"\nendstream")
    offsets = []
    out = io.BytesIO()
    out.write(b"%PDF-1.4\n")
    for i, o in enumerate(objs, start=1):
        offsets.append(out.tell())
        out.write(f"{i} 0 obj\n".encode('latin-1', errors='replace'))
        out.write(o)
        out.write(b"\nendobj\n")
    xref_pos = out.tell()
    out.write(f"xref\n0 {len(objs)+1}\n".encode('latin-1', errors='replace'))
    out.write(b"0000000000 65535 f \n")
    for off in offsets:
        out.write(f"{off:010d} 00000 n \n".encode('latin-1', errors='replace'))
    out.write(b"trailer\n")
    out.write(f"<< /Size {len(objs)+1} /Root 1 0 R >>\n".encode('latin-1', errors='replace'))
    out.write(f"startxref\n{xref_pos}\n%%EOF".encode('latin-1', errors='replace'))
    with open(path, 'wb') as f:
        f.write(out.getvalue())

def _jpeg_size(b):
    try:
        data = b if isinstance(b, (bytes, bytearray)) else b.getvalue()
        i = 0
        while i < len(data) - 1:
            if data[i] == 0xFF and data[i+1] in (0xC0, 0xC2):
                # SOF marker
                if i+7 < len(data):
                    h = (data[i+5] << 8) + data[i+6]
                    w = (data[i+7] << 8) + data[i+8]
                    return w, h
                break
            i += 1
    except Exception:
        pass
    return 640, 360

def _build_basic_pdf_v2(state, season, path):
    all_places = _read_places_csv()
    if season and season != 'All':
        def match(p):
            t = (p.get('best_text') or '')
            return (p.get('best') == season) or (season.lower() in t.lower())
        all_places = [p for p in all_places if match(p)]
    if state:
        all_places = [p for p in all_places if p['state'] == state]
    def rv(x):
        try:
            return float(x['rating']) if x['rating'] else 0.0
        except Exception:
            return 0.0
    places = sorted(all_places, key=lambda x: (rv(x), x['name']), reverse=True)[:12]
    objs = []
    def add_obj_bytes(b): objs.append(b)
    def add_obj_str(s): objs.append(str(s).encode('latin-1', errors='replace'))
    # 1) Font object will be ID 1
    add_obj_str("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    font_id = 1
    # 2) Pages object will be ID 2 with forward references to page IDs
    n = len(places)
    page_ids = [5 + i*3 for i in range(n)]
    kids = " ".join([f"{pid} 0 R" for pid in page_ids]) if n > 0 else ""
    add_obj_str(f"<< /Type /Pages /Count {n} /Kids [{kids}] >>")
    # 3) For each place: add image (ID 3+), content (ID 4+), page (ID 5+)
    for i, p in enumerate(places):
        title = f"{p['name']} · {p['city']}, {p['state']}"
        details = [
            f"Best: {p.get('best') or 'All'}  Rating: {p.get('rating') or '—'}  Fee: {p.get('fee') or '—'}",
            f"Time Needed: {(p.get('time_hrs') and (str(p.get('time_hrs'))+' hrs')) or '—'}  Weekly Off: {p.get('weekly_off') or '—'}",
            f"Season Note: {p.get('best_text') or '—'}",
        ]
        # Image object
        img_name = f"/Im{i+1}"
        img_stream = _fetch_and_compress(p.get('image'))
        w, h = (640, 360)
        img_len = 0
        if img_stream:
            try:
                data = img_stream.getvalue() if hasattr(img_stream, 'getvalue') else img_stream.read()
                img_len = len(data)
                w, h = _jpeg_size(data)
                add_obj_bytes(b"<< /Type /XObject /Subtype /Image /ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /DCTDecode ")
                add_obj_bytes(f"/Width {max(1,w)} /Height {max(1,h)} /Length {img_len} >>\nstream\n".encode('latin-1'))
                add_obj_bytes(data)
                add_obj_bytes(b"\nendstream")
                image_id = len(objs)
            except Exception:
                image_id = None
        else:
            image_id = None
        # Content stream
        def esc(s):
            t = (s or '').replace('–','-').replace('—','-').replace('’',"'").replace('‘',"'").replace('“','\"').replace('”','\"').replace('…','...')
            return t.replace('\\','\\\\').replace('(','\\(').replace(')','\\)')
        lines = ["BT /F1 14 Tf 50 780 Td", f"({esc(title)}) Tj", "0 -24 Td"]
        for d in details:
            lines.append(f"({esc(d)}) Tj 0 -18 Td")
        lines.append("ET")
        if image_id:
            draw_w = 400
            draw_h = int(draw_w * (h/float(w))) if w and h else 250
            lines.append(f"q {draw_w} 0 0 {draw_h} 50 420 cm {img_name} Do Q")
        content = "\n".join(lines).encode('latin-1', errors='replace')
        add_obj_bytes(b"<< /Length " + str(len(content)).encode('latin-1') + b" >>\nstream\n" + content + b"\nendstream")
        content_id = len(objs)
        # Page object referencing Parent 2 0 R
        if image_id:
            page_obj = f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 {font_id} 0 R >> /XObject << {img_name} {image_id} 0 R >> >> /Contents {content_id} 0 R >>"
        else:
            page_obj = f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 {font_id} 0 R >> >> /Contents {content_id} 0 R >>"
        add_obj_str(page_obj)
    # 4) Catalog referencing Pages (object 2)
    add_obj_str("<< /Type /Catalog /Pages 2 0 R >>")
    # Write out
    offsets = []
    out = io.BytesIO()
    out.write(b"%PDF-1.4\n")
    for i, o in enumerate(objs, start=1):
        offsets.append(out.tell())
        out.write(f"{i} 0 obj\n".encode('latin-1', errors='replace'))
        out.write(o)
        out.write(b"\nendobj\n")
    xref = out.tell()
    out.write(f"xref\n0 {len(objs)+1}\n".encode('latin-1', errors='replace'))
    out.write(b"0000000000 65535 f \n")
    for off in offsets:
        out.write(f"{off:010d} 00000 n \n".encode('latin-1', errors='replace'))
    out.write(b"trailer\n")
    out.write(f"<< /Size {len(objs)+1} /Root {len(objs)} 0 R >>\n".encode('latin-1', errors='replace'))
    out.write(f"startxref\n{xref}\n%%EOF".encode('latin-1', errors='replace'))
    with open(path, 'wb') as f:
        f.write(out.getvalue())
@app.route('/download_pdf', methods=['GET'])
@login_required
def download_pdf():
    state = _clean_param(request.args.get('state') or '')
    season = _sanitize_season(request.args.get('season') or 'All')
    force = request.args.get('force') == '1'
    path = _pdf_path(state, season)
    if os.path.exists(path) and not force:
        etag = hashlib.md5((str(os.path.getmtime(path)) + str(os.path.getsize(path))).encode()).hexdigest()
        resp = send_file(path, as_attachment=True, download_name=f"TourEase_{state or 'All'}_{season}.pdf", mimetype='application/pdf', conditional=True)
        resp.headers['Cache-Control'] = 'public, max-age=86400'
        resp.headers['ETag'] = etag
        return resp
    if REPORTLAB_AVAILABLE:
        try:
            _build_pdf(state, season, path)
        except Exception:
            _ensure_pdf_async(state, season, path)
    else:
        try:
            _build_basic_pdf_v2(state, season, path)
        except Exception:
            try:
                _build_basic_pdf(state, season, path)
            except Exception:
                pass
    if os.path.exists(path):
        etag = hashlib.md5((str(os.path.getmtime(path)) + str(os.path.getsize(path))).encode()).hexdigest()
        resp = send_file(path, as_attachment=True, download_name=f"TourEase_{state or 'All'}_{season}.pdf", mimetype='application/pdf', conditional=True)
        resp.headers['Cache-Control'] = 'public, max-age=86400'
        resp.headers['ETag'] = etag
        return resp
    # Final synchronous minimal fallback
    try:
        title = f"TourEase India · {state or 'All States'} · {season}"
        def _sanitize_text(s):
            t = (s or '')
            t = t.replace('–','-').replace('—','-').replace('’',"'").replace('‘',"'").replace('“','"').replace('”','"').replace('…','...').replace('•','-')
            return t
        lines = [_sanitize_text(title), f"Generated: {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')}", "", "No data available yet."]
        # Build an in-memory minimal PDF
        y_start = 800; step = 18
        stream_lines = ["BT /F1 14 Tf 50 800 Td"]
        def esc(s):
            t = _sanitize_text(s)
            return (t or '').replace('\\','\\\\').replace('(','\\(').replace(')','\\)')
        for ln in lines:
            stream_lines.append(f"({esc(ln)}) Tj 0 -{step} Td")
        stream_lines.append("ET")
        sd = ("\n".join(stream_lines)).encode('latin-1', errors='replace')
        objs = [
            b"<< /Type /Catalog /Pages 2 0 R >>",
            b"<< /Type /Pages /Count 1 /Kids [3 0 R] >>",
            b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 5 0 R >> >> /Contents 6 0 R >>",
            b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
            b"<< /Length " + str(len(sd)).encode('latin-1') + b" >>\nstream\n" + sd + b"\nendstream",
        ]
        out = io.BytesIO(); out.write(b"%PDF-1.4\n")
        offsets = [0]
        for i,o in enumerate(objs, start=1):
            offsets.append(out.tell())
            out.write(f"{i} 0 obj\n".encode('latin-1', errors='replace'))
            out.write(o)
            out.write(b"\nendobj\n")
        xref = out.tell()
        out.write(f"xref\n0 {len(objs)+1}\n".encode('latin-1', errors='replace'))
        out.write(b"0000000000 65535 f \n")
        for off in offsets[1:]:
            out.write(f"{off:010d} 00000 n \n".encode('latin-1', errors='replace'))
        out.write(b"trailer\n")
        out.write(f"<< /Size {len(objs)+1} /Root 1 0 R >>\n".encode('latin-1', errors='replace'))
        out.write(f"startxref\n{xref}\n%%EOF".encode('latin-1', errors='replace'))
        out.seek(0)
    except Exception:
        return jsonify({'error':'Failed to generate PDF'}), 500
@app.route('/api/destinations', methods=['GET'])
def api_destinations():
    db = get_db()
    rows = db.execute("SELECT * FROM destinations WHERE name != 'Shimla' ORDER BY name").fetchall()
    def image_for(name):
        n = name or ''
        if 'Goa' in n: return "https://upload.wikimedia.org/wikipedia/commons/thumb/b/b1/Baga_Beach.jpg/640px-Baga_Beach.jpg"
        if 'Hyderabad' in n: return "https://upload.wikimedia.org/wikipedia/commons/thumb/d/d9/Charminar_%28Night_View%29.jpg/640px-Charminar_%28Night_View%29.jpg"
        if 'Delhi' in n: return "https://upload.wikimedia.org/wikipedia/commons/thumb/1/19/Inscription_on_India_Gate.jpg/640px-Inscription_on_India_Gate.jpg"
        if 'Munnar' in n: return "https://upload.wikimedia.org/wikipedia/commons/thumb/7/71/Munnar_tea_estate.jpg/640px-Munnar_tea_estate.jpg"
        return "https://upload.wikimedia.org/wikipedia/commons/6/6e/India_collage.jpg"
    out = []
    for r in rows:
        out.append({
            'id': r['id'],
            'name': r['name'],
            'state': r['state'],
            'best_season': r['best_season'],
            'description': r['description'],
            'map_embed_url': r['map_embed_url'],
            'image_url': image_for(r['name']),
            'lat': r['lat'],
            'lng': r['lng'],
        })
    return jsonify({'destinations': out})
@app.route('/api/login', methods=['POST'])
def api_login():
    init_firebase()
    if not FIREBASE_ENABLED:
        return jsonify({'error': 'Firebase not configured'}), 500
    data = request.get_json(force=True)
    token = data.get('idToken')
    if not token:
        return jsonify({'error': 'Missing idToken'}), 400
    try:
        decoded = fb_auth.verify_id_token(token)
        uid = decoded.get('uid')
        email = decoded.get('email')
        name = decoded.get('name')
        phone = decoded.get('phone_number')
        session['uid'] = uid
        session['user_email'] = email
        session['user_name'] = name or email
        FS.collection('users').document(uid).set({
            'uid': uid,
            'name': name,
            'email': email,
            'phone': phone,
            'isOnline': True,
            'registered_at': firestore.SERVER_TIMESTAMP
        }, merge=True)
        act_ref = FS.collection('user_activity').document()
        act_ref.set({
            'uid': uid,
            'login_ts': datetime.utcnow(),
            'logout_ts': None
        })
        session['activity_id'] = act_ref.id
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': 'Invalid token'}), 401

@app.route('/api/stats')
def api_stats():
    init_firebase()
    if not FIREBASE_ENABLED:
        return jsonify({'total': TOTAL_USERS, 'active': ACTIVE_USERS})
    try:
        users = list(FS.collection('users').stream())
        active = list(FS.collection('users').where('isOnline', '==', True).stream())
        return jsonify({'total': len(users), 'active': len(active)})
    except Exception:
        return jsonify({'total': TOTAL_USERS, 'active': ACTIVE_USERS})

@app.route('/api/simple_login', methods=['POST'])
def api_simple_login():
    data = request.get_json(force=True)
    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip()
    phone = (data.get('phone') or '').strip()
    if not email:
        return jsonify({'error': 'Missing email'}), 400
    uid = uuid.uuid4().hex
    session['uid'] = uid
    session['user_email'] = email
    session['user_name'] = name or email
    session['user_phone'] = phone
    global ACTIVE_USERS, TOTAL_USERS
    ACTIVE_USERS += 1
    TOTAL_USERS += 1
    try:
        init_firebase()
        if FIREBASE_ENABLED:
            FS.collection('users').document(uid).set({
                'uid': uid,
                'name': name or email,
                'email': email,
                'phone': phone,
                'isOnline': True,
                'registered_at': firestore.SERVER_TIMESTAMP
            }, merge=True)
            act_ref = FS.collection('user_activity').document()
            act_ref.set({
                'uid': uid,
                'login_ts': datetime.utcnow(),
                'logout_ts': None
            })
            session['activity_id'] = act_ref.id
    except Exception:
        pass
    return jsonify({'ok': True})

@app.route('/api/favorites/add', methods=['POST'])
@login_required
def add_favorite():
    """Add a place to user favorites"""
    try:
        data = request.get_json(force=True)
        uid = session.get('uid')
        place_name = (data.get('place_name') or '').strip()
        city = (data.get('city') or '').strip()
        state = (data.get('state') or '').strip()
        image_url = (data.get('image_url') or '').strip()
        rating = (data.get('rating') or '').strip()
        
        if not (place_name and state and uid):
            return jsonify({'error': 'Missing required fields'}), 400
        
        db = get_db()
        try:
            db.execute(
                "INSERT INTO favorites (user_id, place_name, city, state, image_url, rating) VALUES (?, ?, ?, ?, ?, ?)",
                (uid, place_name, city, state, image_url, rating)
            )
            db.commit()
            return jsonify({'ok': True, 'message': 'Added to favorites ❤️'})
        except Exception as e:
            if 'UNIQUE constraint failed' in str(e):
                # Remove if already exists (toggle)
                db.execute(
                    "DELETE FROM favorites WHERE user_id=? AND place_name=? AND state=?",
                    (uid, place_name, state)
                )
                db.commit()
                return jsonify({'ok': True, 'message': 'Removed from favorites', 'removed': True})
            raise
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/favorites/remove', methods=['POST'])
@login_required
def remove_favorite():
    """Remove a place from user favorites"""
    try:
        data = request.get_json(force=True)
        uid = session.get('uid')
        place_name = (data.get('place_name') or '').strip()
        state = (data.get('state') or '').strip()
        
        if not (place_name and state and uid):
            return jsonify({'error': 'Missing required fields'}), 400
        
        db = get_db()
        db.execute(
            "DELETE FROM favorites WHERE user_id=? AND place_name=? AND state=?",
            (uid, place_name, state)
        )
        db.commit()
        return jsonify({'ok': True, 'message': 'Removed from favorites'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/favorites/list')
@login_required
def get_favorites():
    """Get user's favorite places"""
    try:
        uid = session.get('uid')
        db = get_db()
        favs = db.execute(
            "SELECT id, place_name, city, state, image_url, rating, added_at FROM favorites WHERE user_id=? ORDER BY added_at DESC LIMIT 100",
            (uid,)
        ).fetchall()
        result = [{
            'id': f['id'],
            'place_name': f['place_name'],
            'city': f['city'],
            'state': f['state'],
            'image_url': f['image_url'],
            'rating': f['rating'],
            'added_at': f['added_at']
        } for f in favs]
        return jsonify({'favorites': result, 'count': len(result)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/favorites/check', methods=['GET'])
@login_required
def check_favorite():
    """Check if a place is in user favorites"""
    try:
        uid = session.get('uid')
        place_name = request.args.get('name', '').strip()
        state = request.args.get('state', '').strip()
        
        if not (place_name and state):
            return jsonify({'is_favorite': False})
        
        db = get_db()
        result = db.execute(
            "SELECT id FROM favorites WHERE user_id=? AND place_name=? AND state=?",
            (uid, place_name, state)
        ).fetchone()
        
        return jsonify({'is_favorite': bool(result)})
    except Exception as e:
        return jsonify({'is_favorite': False})

@app.route('/favorites')
@login_required
def view_favorites():
    """View user's saved favorites page"""
    db = get_db()
    uid = session.get('uid')
    favs = db.execute(
        "SELECT place_name, city, state, image_url, rating FROM favorites WHERE user_id=? ORDER BY added_at DESC",
        (uid,)
    ).fetchall()
    return render_template('favorites.html', favorites=favs, count=len(favs))

if __name__ == '__main__':
    # Ensure DB and tables exist; if packages table missing, run schema
    create_needed = False
    if not os.path.exists(DATABASE):
        create_needed = True
    else:
        try:
            db = sqlite3.connect(DATABASE)
            cur = db.cursor()
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='packages';")
            row = cur.fetchone()
            if not row:
                create_needed = True
            db.close()
        except Exception:
            create_needed = True
    if create_needed:
        with app.app_context():
            init_db()
            ensure_columns()
    else:
        with app.app_context():
            ensure_columns()
    init_firebase()
    def prewarm():
        try:
            with app.app_context():
                all_places = _read_places_csv()
                states = sorted({p['state'] for p in all_places if p['state']})
                seasons = ['All','Winter','Monsoon','Summer']
                for s in states[:20]:
                    for se in seasons:
                        try:
                            _build_pdf(s, se, _pdf_path(s, se))
                        except Exception:
                            pass
        except Exception:
            pass
    threading.Thread(target=prewarm, daemon=True).start()
    app.run(debug=True)
